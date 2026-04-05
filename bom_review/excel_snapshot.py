"""Excel 범위 선택 후 타임스탬프 결과 통합문서에 복사하고 Range_Set 시트를 유지한다."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell

from bom_review.bom_parse import normalize_designators_to_comma_space
from bom_review.excel_com import (
    SelectionSourceMeta,
    copy_worksheet_to_workbook_end,
    excel_a1_address_bounds,
    format_com_error,
)

RANGE_SET_SHEET = "Range_Set"
# 결과 통합문서에서만 BOM 좌표명(', ' 통일) 후처리 대상이 되는 시트명(원본 파일·원본_검토복사 등은 변경하지 않음).
BOM_REVIEW_COPY_SHEET = "BOM_검토복사"
# 시트UsedRange·검토범위 값은 결과 통합문서의 «복사시트» 기준 주소이다.
RANGE_SET_HEADERS = (
    "역할",
    "원본파일",
    "원본시트",
    "결과시트UsedRange",
    "결과검토범위",
    "복사시트",
)


def new_snapshot_workbook_path(folder: Path) -> Path:
    """작업 폴더에 저장할 결과 통합문서 경로(타임스탬프 파일명)."""
    folder = folder.resolve()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return folder / f"BOM_Review_Result_{ts}.xlsx"


def destination_sheet_name_for_role(role: str) -> str:
    """역할별 복사 대상 시트 이름(Excel 제한 내)."""
    if role == "BOM":
        return BOM_REVIEW_COPY_SHEET
    if role == "원본":
        return "원본_검토복사"
    base = (
        role.replace("[", "")
        .replace("]", "")
        .replace("*", "")
        .replace("?", "")
        .replace("/", "")
        .replace("\\", "")[:28]
    )
    return base or "데이터"


def _ensure_range_set_sheet(wb: Any) -> Any:
    """Range_Set 시트가 없으면 맨 앞에 만든다."""
    if RANGE_SET_SHEET in wb.sheetnames:
        return wb[RANGE_SET_SHEET]
    ws = wb.create_sheet(RANGE_SET_SHEET, 0)
    ws.append(list(RANGE_SET_HEADERS))
    return ws


def _upsert_range_set_row(wb: Any, role: str, meta: SelectionSourceMeta, copy_sheet: str) -> None:
    """동일 역할 행은 제거한 뒤 새 메타 한 줄을 추가한다."""
    rs = _ensure_range_set_sheet(wb)
    ncols = len(RANGE_SET_HEADERS)
    kept: list[list[Any]] = []
    for row in rs.iter_rows(min_row=2, values_only=True):
        if not row or all(c is None for c in row):
            continue
        r0 = row[0]
        if r0 is not None and str(r0).strip() == role:
            continue
        kept.append([row[i] if i < len(row) else None for i in range(ncols)])
    if rs.max_row >= 2:
        rs.delete_rows(2, rs.max_row - 1)
    for r in kept:
        rs.append(r)
    rs.append(
        [
            role,
            meta.source_file,
            meta.source_sheet,
            meta.sheet_used_range_address,
            meta.review_range_address,
            copy_sheet,
        ]
    )


def _create_minimal_snapshot_workbook(snapshot_path: Path) -> None:
    """COM 복사 전에 쓸 빈 결과 통합문서(Range_Set만)."""
    wb = Workbook()
    ws = wb.active
    ws.title = RANGE_SET_SHEET
    ws.append(list(RANGE_SET_HEADERS))
    snapshot_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(snapshot_path)
    wb.close()


def _normalize_bom_coord_in_table(
    headers: list[str],
    data_rows: list[list[Any]],
    matrix_col_index: int,
) -> tuple[list[str], list[list[Any]]]:
    """UsedRange 기준 열 인덱스(0-based)의 «데이터 행» 좌표명 셀만 ', ' 형태로 통일(헤더 문자열은 그대로)."""
    hdr = list(headers)
    dat = [list(r) for r in data_rows]
    idx = matrix_col_index
    for r in dat:
        if idx < len(r):
            r[idx] = normalize_designators_to_comma_space(r[idx])
    return hdr, dat


def finalize_snapshot_openpyxl(
    snapshot_path: Path,
    *,
    role: str,
    dest_sheet_name: str,
    meta: SelectionSourceMeta,
    bom_coord_excel_col_1based: int | None,
    bom_norm_row_start: int,
    bom_norm_row_end: int,
    bom_first_data_row_1based: int | None = None,
    write_range_set_row: bool = True,
) -> None:
    """COM 저장 후 결과 파일의 BOM 복사 시트 좌표 열만 정규화하고 Range_Set 반영(openpyxl, 단일 저장).

    bom_first_data_row_1based: 검토 범위 첫 행(r1)이 열 헤더이므로, 좌표 정규화는 r1+1행부터만 적용.
    None이면 bom_norm_row_start+1(UsedRange 첫 행만 헤더로 가정)으로 동작한다.

    write_range_set_row: False면 Range_Set 행을 넣지 않는다(Excel 1단계 시트 복사만 한 직후 등).
    """
    snapshot_path = snapshot_path.resolve()
    wb = load_workbook(snapshot_path)
    try:
        # 원본 통합문서는 COM에서 저장 없이 닫힘. 여기서는 스냅샷 xlsx만 열리며,
        # BOM이 아닌 역할·다른 시트(예: 원본_검토복사)는 좌표 열을 건드리지 않는다.
        normalize_bom_only = (
            role == "BOM"
            and dest_sheet_name == BOM_REVIEW_COPY_SHEET
            and bom_coord_excel_col_1based is not None
        )
        if normalize_bom_only:
            ws = wb[dest_sheet_name]
            c = bom_coord_excel_col_1based
            # COM UsedRange 행 수와 openpyxl max_row가 어긋나면 아래쪽 행이 빠지지 않게 한다.
            last_r = ws.max_row or 1
            end_r = max(bom_norm_row_end, last_r)
            first_data_row = (
                bom_first_data_row_1based
                if bom_first_data_row_1based is not None
                else bom_norm_row_start + 1
            )
            for r in range(bom_norm_row_start, end_r + 1):
                if r < first_data_row:
                    continue
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = normalize_designators_to_comma_space(cell.value)
        _ensure_range_set_sheet(wb)
        if write_range_set_row:
            _upsert_range_set_row(wb, role, meta, dest_sheet_name)
        wb.save(snapshot_path)
    finally:
        wb.close()


ExcelParsedSelection = tuple[
    list[str],
    list[list[Any]],
    list[str],
    list[list[Any]],
    SelectionSourceMeta,
    int,
    int,
    int,
    int,
    int,
    int,
]


def apply_review_selection_to_snapshot(
    snapshot_path: Path,
    *,
    role: str,
    dest_sheet_name: str,
    source_meta: SelectionSourceMeta,
    parsed: ExcelParsedSelection,
    bom_coord_excel_col_1based: int | None = None,
) -> SelectionSourceMeta:
    """
    2단계: 결과 통합문서의 복사 시트에서 잡은 Selection(parsed)으로 Range_Set·BOM 좌표 정규화만 반영한다.
    source_meta의 원본파일·원본시트는 1단계 복사 시 확정된 값을 유지한다.
    """
    (
        full_h,
        full_d,
        _rev_h,
        _rev_d,
        _meta_discard,
        ur_r,
        ur_c,
        r1,
        c1,
        r2,
        c2,
    ) = parsed
    nrow = 1 + len(full_d)
    w = len(full_h)
    used_addr = excel_a1_address_bounds(ur_r, ur_c, ur_r + nrow - 1, ur_c + max(w, 1) - 1)
    rev_addr = excel_a1_address_bounds(r1, c1, r2, c2)
    merged = SelectionSourceMeta(
        source_file=source_meta.source_file,
        source_sheet=source_meta.source_sheet,
        sheet_used_range_address=used_addr,
        review_range_address=rev_addr,
    )
    row_start = ur_r
    row_end = ur_r + len(full_d)
    bom_excel_col: int | None = (
        bom_coord_excel_col_1based if role == "BOM" else None
    )
    finalize_snapshot_openpyxl(
        snapshot_path,
        role=role,
        dest_sheet_name=dest_sheet_name,
        meta=merged,
        bom_coord_excel_col_1based=bom_excel_col,
        bom_norm_row_start=row_start,
        bom_norm_row_end=row_end,
        bom_first_data_row_1based=(r1 + 1) if role == "BOM" else None,
        write_range_set_row=True,
    )
    return merged


def persist_role_sheet_via_com(
    xl: Any,
    source_wb: Any,
    snapshot_path: Path,
    role: str,
    parsed: ExcelParsedSelection,
    *,
    bom_coord_excel_col_1based: int | None = None,
    defer_openpyxl_finalize: bool = False,
) -> tuple[SelectionSourceMeta, str]:
    """
    원본 시트 전체를 Excel COM으로 결과 통합문서에 복사(서식 유지)하고,
    원본 통합문서를 닫은 뒤 Range_Set·BOM 좌표 정규화를 openpyxl로 마친다.

    bom_coord_excel_col_1based: role이 BOM일 때만, 좌표명 열의 시트 절대 열 번호(1-based).

    defer_openpyxl_finalize: True면 시트 복사·저장만 하고 Range_Set·openpyxl 후처리는 하지 않는다.
    (GUI 2단계에서 범위 확정 후 apply_review_selection_to_snapshot이 기록한다.)

    SelectionSourceMeta의 시트UsedRange·검토범위 주소는 결과 파일의 복사 시트 기준이다.
    """
    (
        full_h,
        full_d,
        _rev_h,
        _rev_d,
        meta_src,
        ur_r,
        ur_c,
        r1,
        c1,
        r2,
        c2,
    ) = parsed
    dest_name = destination_sheet_name_for_role(role)
    snapshot_path = snapshot_path.resolve()
    snapshot_path.parent.mkdir(parents=True, exist_ok=True)
    if not snapshot_path.exists():
        _create_minimal_snapshot_workbook(snapshot_path)

    dest_wb: Any | None = None
    meta: SelectionSourceMeta
    copied_ur_r: int | None = None
    xl.DisplayAlerts = False
    try:
        # 결과 통합문서를 연 뒤에는 ActiveWorkbook이 바뀌므로, 그 전에 원본 시트 참조를 잡는다.
        sk = (meta_src.source_sheet or "").strip()
        if sk and sk != "?":
            try:
                src_ws = source_wb.Worksheets(sk)
            except Exception:  # noqa: BLE001
                src_ws = source_wb.ActiveSheet
        else:
            src_ws = source_wb.ActiveSheet
        actual_sheet = str(src_ws.Name)

        def _source_file_label(wb: Any) -> str:
            try:
                return Path(str(wb.FullName)).name
            except Exception:  # noqa: BLE001
                return str(wb.Name)

        src_file_label = (
            meta_src.source_file
            if (meta_src.source_file or "").strip() not in ("", "?")
            else _source_file_label(source_wb)
        )

        dest_wb = xl.Workbooks.Open(str(snapshot_path))
        for i in range(int(dest_wb.Worksheets.Count), 0, -1):
            sh = dest_wb.Worksheets(i)
            if str(sh.Name) == dest_name:
                sh.Delete()
        new_ws = copy_worksheet_to_workbook_end(src_ws, dest_wb)
        new_ws.Name = dest_name
        ur = new_ws.UsedRange
        ur_row = int(ur.Row)
        ur_col = int(ur.Column)
        copied_ur_r = ur_row
        used_addr = excel_a1_address_bounds(
            ur_row,
            ur_col,
            ur_row + int(ur.Rows.Count) - 1,
            ur_col + int(ur.Columns.Count) - 1,
        )
        rev_addr = excel_a1_address_bounds(r1, c1, r2, c2)
        meta = SelectionSourceMeta(
            source_file=src_file_label,
            source_sheet=actual_sheet,
            sheet_used_range_address=used_addr,
            review_range_address=rev_addr,
        )
        dest_wb.Save()
        dest_wb.Close(SaveChanges=False)
        dest_wb = None
        source_wb.Close(SaveChanges=False)
    except Exception as e:
        if dest_wb is not None:
            try:
                dest_wb.Close(SaveChanges=False)
            except Exception:  # noqa: BLE001
                pass
        try:
            source_wb.Close(SaveChanges=False)
        except Exception:  # noqa: BLE001
            pass
        try:
            xl.Quit()
        except Exception:  # noqa: BLE001
            pass
        try:
            from pywintypes import com_error as ComError
        except ImportError:
            pass
        else:
            if isinstance(e, ComError):
                raise RuntimeError(
                    "검토용 통합문서(Excel COM) 오류: " + format_com_error(e)
                ) from e
        raise
    finally:
        try:
            xl.DisplayAlerts = True
        except Exception:  # noqa: BLE001
            pass
    try:
        xl.Quit()
    except Exception:  # noqa: BLE001
        pass

    if defer_openpyxl_finalize:
        return meta, dest_name

    bom_excel_col: int | None = (
        bom_coord_excel_col_1based if role == "BOM" else None
    )
    row_start = copied_ur_r if copied_ur_r is not None else ur_r
    row_end = row_start + len(full_d)
    finalize_snapshot_openpyxl(
        snapshot_path,
        role=role,
        dest_sheet_name=dest_name,
        meta=meta,
        bom_coord_excel_col_1based=bom_excel_col,
        bom_norm_row_start=row_start,
        bom_norm_row_end=row_end,
        bom_first_data_row_1based=(r1 + 1) if role == "BOM" else None,
    )
    return meta, dest_name


def write_role_range_to_snapshot(
    snapshot_path: Path,
    *,
    role: str,
    headers: list[str],
    data_rows: list[list[Any]],
    meta: SelectionSourceMeta,
    create_new_workbook: bool,
    bom_coord_matrix_col_index: int | None = None,
) -> str:
    """
    결과 통합문서에 헤더+데이터 시트를 쓰고 Range_Set을 갱신한다.
    (테스트·비-COM 경로용. GUI Excel 적용은 persist_role_sheet_via_com 사용.)

    role이 BOM이고 bom_coord_matrix_col_index가 있으면 해당 열(시트 UsedRange 기준)만 정규화한다.
    """
    snapshot_path = snapshot_path.resolve()
    dest_name = destination_sheet_name_for_role(role)

    hdr, rows = headers, data_rows
    if role == "BOM" and bom_coord_matrix_col_index is not None:
        hdr, rows = _normalize_bom_coord_in_table(headers, data_rows, bom_coord_matrix_col_index)

    if snapshot_path.exists():
        wb = load_workbook(snapshot_path)
    else:
        if not create_new_workbook:
            raise FileNotFoundError(f"결과 통합문서가 없습니다: {snapshot_path}")
        wb = Workbook()
        default = wb.active
        default.title = RANGE_SET_SHEET
        default.append(list(RANGE_SET_HEADERS))

    _ensure_range_set_sheet(wb)

    if dest_name in wb.sheetnames:
        wb.remove(wb[dest_name])
    ws_data = wb.create_sheet(dest_name)
    ws_data.append(list(hdr))
    for row in rows:
        ws_data.append(list(row))

    _upsert_range_set_row(wb, role, meta, dest_name)

    snapshot_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(snapshot_path)
    wb.close()
    return dest_name

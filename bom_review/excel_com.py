"""
Excel COM (Windows) — 통합문서를 띄우고 Selection 값을 2차원으로 읽는다.
pywin32 필요: pip install pywin32
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

EXCEL_SUFFIXES = frozenset({".xlsx", ".xlsm", ".xlsb", ".xls"})


@dataclass(frozen=True)
class SelectionSourceMeta:
    """
    Range_Set용 메타.
    원본파일·원본시트: 출처.
    sheet_used_range_address·review_range_address: 결과 통합문서의 «복사시트» 기준 절대 주소.
    """

    source_file: str
    source_sheet: str
    sheet_used_range_address: str
    review_range_address: str


def is_excel_path(path: Path) -> bool:
    return path.suffix.lower() in EXCEL_SUFFIXES


def excel_a1_address_bounds(r1: int, c1: int, r2: int, c2: int) -> str:
    """
    Excel 1-based 행·열로 절대 참조 문자열 ($A$1:$B$2).

    pywin32 late binding에서 Range.Address 가 문자열로만 잡혀
    Address(RowAbsolute=...) 호출 시 'str' object is not callable 이 나는 경우가 있어,
    COM Address 호출 없이 주소를 만든다.
    """
    from openpyxl.utils import get_column_letter

    return f"${get_column_letter(c1)}${r1}:${get_column_letter(c2)}${r2}"


def normalize_com_value(val: Any) -> list[list[Any]]:
    """Excel Range.Value → 균일한 2차원 리스트 (빈 셀은 None)."""
    if val is None:
        return []
    if not isinstance(val, tuple):
        return [[val]]
    if val and not any(isinstance(x, tuple) for x in val):
        return [list(val)]
    rows: list[list[Any]] = []
    for row in val:
        if isinstance(row, tuple):
            rows.append(list(row))
        else:
            rows.append([row])
    return rows


def _pad_rows(rows: list[list[Any]], width: int) -> list[list[Any]]:
    out: list[list[Any]] = []
    for r in rows:
        row = list(r)
        if len(row) < width:
            row.extend([None] * (width - len(row)))
        out.append(row[:width])
    return out


def _row_to_headers(cells: list[Any]) -> list[str]:
    return [
        str(c).strip() if c is not None and str(c).strip() else f"열{i}"
        for i, c in enumerate(cells)
    ]


def read_selection_as_header_and_rows(xl: Any) -> tuple[list[str], list[list[Any]]] | None:
    """
    현재 Excel Selection → (헤더명 목록, 데이터 행들).
    첫 행을 헤더로 쓴다. 선택이 없거나 비어 있으면 None.
    """
    try:
        sel = xl.Selection
        val = sel.Value
    except Exception:  # noqa: BLE001
        return None
    rows = normalize_com_value(val)
    if not rows:
        return None
    w = max(len(r) for r in rows)
    rows = _pad_rows(rows, w)
    headers = _row_to_headers(rows[0])
    data = rows[1:]
    if data:
        data = _pad_rows(data, len(headers))
    return headers, data


def read_full_sheet_and_review_selection(
    xl: Any,
) -> tuple[
    list[str],
    list[list[Any]],
    list[str],
    list[list[Any]],
    SelectionSourceMeta,
    int,
    int,
    int,
    int,
    int,
    int,
] | None:
    """
    활성 시트 UsedRange 전체(복사본) + 현재 Selection(검토·콤보용)을 동시에 읽는다.

    반환: (full_headers, full_data_rows, review_headers, review_data_rows, meta,
           ur_row, ur_col, sel_r1, sel_c1, sel_r2, sel_c2) — 둘 다 Excel 1-based.
    Selection은 UsedRange 안에 있어야 한다.
    """
    try:
        sel = xl.Selection
        ws = sel.Worksheet
        ur = ws.UsedRange
        ur_val = ur.Value
    except Exception:  # noqa: BLE001
        return None
    full_matrix = normalize_com_value(ur_val)
    if not full_matrix:
        return None
    w = max(len(r) for r in full_matrix)
    full_matrix = _pad_rows(full_matrix, w)
    ur_r = int(ur.Row)
    ur_c = int(ur.Column)

    full_headers = _row_to_headers(full_matrix[0])
    full_data = full_matrix[1:]
    if full_data:
        full_data = _pad_rows(full_data, len(full_headers))

    r1, c1 = int(sel.Row), int(sel.Column)
    r2 = r1 + int(sel.Rows.Count) - 1
    c2 = c1 + int(sel.Columns.Count) - 1

    ir1, ic1 = r1 - ur_r, c1 - ur_c
    ir2, ic2 = r2 - ur_r, c2 - ur_c
    if ir1 < 0 or ic1 < 0 or ir2 >= len(full_matrix) or ic2 >= w:
        return None

    rev_header_cells = full_matrix[ir1][ic1 : ic2 + 1]
    review_headers = _row_to_headers(rev_header_cells)
    review_data: list[list[Any]] = []
    for ir in range(ir1 + 1, ir2 + 1):
        row = full_matrix[ir]
        chunk = list(row[ic1 : ic2 + 1])
        if len(chunk) < len(review_headers):
            chunk.extend([None] * (len(review_headers) - len(chunk)))
        review_data.append(chunk[: len(review_headers)])

    # 시트 이름·파일명은 Address() 실패와 무관하게 확보 (Address만 실패해도 '?' 시트로 복사 시도하지 않음)
    sheet_nm = str(ws.Name)
    parent_wb = ws.Parent
    try:
        src_name = Path(str(parent_wb.FullName)).name
    except Exception:  # noqa: BLE001
        try:
            src_name = str(parent_wb.Name)
        except Exception:  # noqa: BLE001
            src_name = "?"

    r_end_used = ur_r + len(full_matrix) - 1
    c_end_used = ur_c + w - 1
    used_addr = excel_a1_address_bounds(ur_r, ur_c, r_end_used, c_end_used)
    rev_addr = excel_a1_address_bounds(r1, c1, r2, c2)

    meta = SelectionSourceMeta(
        source_file=src_name,
        source_sheet=sheet_nm,
        sheet_used_range_address=used_addr,
        review_range_address=rev_addr,
    )

    return (
        full_headers,
        full_data,
        review_headers,
        review_data,
        meta,
        ur_r,
        ur_c,
        r1,
        c1,
        r2,
        c2,
    )


def read_active_sheet_full_used_as_selection(
    xl: Any,
) -> tuple[
    list[str],
    list[list[Any]],
    list[str],
    list[list[Any]],
    SelectionSourceMeta,
    int,
    int,
    int,
    int,
    int,
    int,
] | None:
    """
    활성 시트(사용자가 선택한 시트) UsedRange 전체를 검토 범위와 동일하게 잡아 읽는다.
    1단계 «시트만 복사» 시 Selection 없이 복사 대상 시트 내용을 확정할 때 쓴다.
    """
    try:
        ws = xl.ActiveSheet
        ur = ws.UsedRange
        ur_val = ur.Value
    except Exception:  # noqa: BLE001
        return None
    full_matrix = normalize_com_value(ur_val)
    if not full_matrix:
        return None
    w = max(len(r) for r in full_matrix)
    full_matrix = _pad_rows(full_matrix, w)
    ur_r = int(ur.Row)
    ur_c = int(ur.Column)

    full_headers = _row_to_headers(full_matrix[0])
    full_data = full_matrix[1:]
    if full_data:
        full_data = _pad_rows(full_data, len(full_headers))

    nrows = len(full_matrix)
    r1, c1 = ur_r, ur_c
    r2, c2 = ur_r + nrows - 1, ur_c + w - 1

    review_headers = list(full_headers)
    review_data = [list(r) for r in full_data]

    sheet_nm = str(ws.Name)
    parent_wb = ws.Parent
    try:
        src_name = Path(str(parent_wb.FullName)).name
    except Exception:  # noqa: BLE001
        try:
            src_name = str(parent_wb.Name)
        except Exception:  # noqa: BLE001
            src_name = "?"

    r_end_used = ur_r + nrows - 1
    c_end_used = ur_c + w - 1
    used_addr = excel_a1_address_bounds(ur_r, ur_c, r_end_used, c_end_used)
    rev_addr = used_addr

    meta = SelectionSourceMeta(
        source_file=src_name,
        source_sheet=sheet_nm,
        sheet_used_range_address=used_addr,
        review_range_address=rev_addr,
    )

    return (
        full_headers,
        full_data,
        review_headers,
        review_data,
        meta,
        ur_r,
        ur_c,
        r1,
        c1,
        r2,
        c2,
    )


def format_com_error(exc: BaseException) -> str:
    """pywin32 com_error 등 COM 예외를 사용자 안내용 문자열로 요약한다."""
    try:
        from pywintypes import com_error as ComError
    except ImportError:
        return str(exc)
    if isinstance(exc, ComError):
        a = exc.args
        parts: list[str] = []
        if len(a) > 0:
            parts.append(f"hresult={a[0]!r}")
        if len(a) > 1:
            parts.append(f"텍스트={a[1]!r}")
        if len(a) > 2 and a[2] is not None:
            parts.append(f"정보={a[2]!r}")
        return "COM " + ", ".join(parts) if parts else repr(exc)
    return str(exc)


def copy_worksheet_to_workbook_end(src_ws: Any, dest_wb: Any) -> Any:
    """
    src_ws를 dest_wb의 마지막 시트 뒤에 복사한 뒤, 새로 생긴 시트(마지막)를 반환.

    Worksheet.Copy(After=...) 만 쓰면 일부 Excel·pywin32 조합에서
    DISP_E_BADPARAMCOUNT(-2147352565) 가 날 수 있어, VBA 시그니처
    Copy(Before, After)에 맞춰 위치 인자(None / Missing, after)를 우선 시도한다.
    """
    import pythoncom
    from pywintypes import com_error as ComError

    n = int(dest_wb.Worksheets.Count)
    after_sh = dest_wb.Worksheets(n)

    def _try_copy() -> None:
        last_inner: BaseException | None = None
        for fn in (
            lambda: src_ws.Copy(None, after_sh),
            lambda: src_ws.Copy(pythoncom.Missing, after_sh),
            lambda: src_ws.Copy(After=after_sh),
        ):
            try:
                fn()
                return
            except Exception as e:  # noqa: BLE001
                last_inner = e
        assert last_inner is not None
        raise last_inner

    try:
        _try_copy()
    except Exception as e:  # noqa: BLE001
        if isinstance(e, ComError):
            raise RuntimeError(
                "시트 복사(Copy) 단계에서 Excel COM 오류: " + format_com_error(e)
            ) from e
        raise RuntimeError(f"시트 복사(Copy) 실패: {e}") from e

    return dest_wb.Worksheets(int(dest_wb.Worksheets.Count))


def open_workbook_in_new_excel(path: Path) -> tuple[Any, Any]:
    """새 Excel 인스턴스를 띄우고 통합문서를 연다. (xl, wb)"""
    try:
        from win32com.client import DispatchEx
    except ImportError as e:
        raise RuntimeError(
            "Excel 연동에 pywin32 가 필요합니다. 다음을 실행하세요: pip install pywin32"
        ) from e
    path = path.resolve()
    xl = DispatchEx("Excel.Application")
    xl.Visible = True
    wb = xl.Workbooks.Open(str(path), UpdateLinks=0, ReadOnly=False)
    return xl, wb


def close_excel_quietly(xl: Any, rw: Any) -> None:
    """저장 없이 통합문서 닫기 및 Excel 종료."""
    try:
        if rw is not None:
            rw.Close(SaveChanges=False)
    except Exception:  # noqa: BLE001
        pass
    try:
        if xl is not None:
            xl.Quit()
    except Exception:  # noqa: BLE001
        pass

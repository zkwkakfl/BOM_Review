"""CSV / Excel(xlsx·xlsm·xlsb) 테이블 읽기 — 첫 시트·헤더 행 기준."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

__all__ = [
    "peek_headers",
    "load_header_and_rows",
    "load_header_and_rows_by_sheet_name",
    "resolve_column_index",
    "values_for_column",
    "supported_extensions",
]


def supported_extensions() -> frozenset[str]:
    return frozenset({".csv", ".xlsx", ".xlsm", ".xlsb"})


def _cell_str(v: Any, i: int) -> str:
    if v is None:
        return f"열{i}"
    return str(v).strip() or f"열{i}"


def peek_headers(path: Path, *, sheet_index: int = 0) -> list[str]:
    """첫 행을 헤더로 간주한 열 이름 목록."""
    headers, _ = load_header_and_rows(path, sheet_index=sheet_index, max_data_rows=0)
    return headers


def load_header_and_rows(
    path: Path,
    *,
    sheet_index: int = 0,
    max_data_rows: int | None = None,
) -> tuple[list[str], list[list[Any]]]:
    """
    (headers, data_rows). data_rows는 헤더 다음부터.
    max_data_rows가 0이면 헤더만 (peek용), None이면 전부.
    """
    ext = path.suffix.lower()
    if ext == ".csv":
        return _load_csv(path, max_data_rows=max_data_rows)
    if ext in (".xlsx", ".xlsm", ".xlsb"):
        return _load_xlsx(path, sheet_index=sheet_index, max_data_rows=max_data_rows)
    raise ValueError(f"지원하지 않는 확장자입니다: {ext}")


def load_header_and_rows_by_sheet_name(
    path: Path,
    *,
    sheet_name: str,
    max_data_rows: int | None = None,
) -> tuple[list[str], list[list[Any]]]:
    """이름으로 시트를 고른 뒤 (headers, data_rows). xlsx·xlsm·xlsb만."""
    ext = path.suffix.lower()
    if ext not in (".xlsx", ".xlsm", ".xlsb"):
        raise ValueError(f"시트 이름으로 읽기는 Excel 통합문서만 지원합니다: {ext}")
    return _load_xlsx_by_sheet_name(path, sheet_name=sheet_name, max_data_rows=max_data_rows)


def _load_csv(
    path: Path,
    *,
    max_data_rows: int | None,
) -> tuple[list[str], list[list[Any]]]:
    last_err: Exception | None = None
    for enc in ("utf-8-sig", "utf-8", "cp949"):
        try:
            with path.open(newline="", encoding=enc) as f:
                reader = csv.reader(f)
                try:
                    header_row = next(reader)
                except StopIteration:
                    return [], []
                headers = [_cell_str(c, i) for i, c in enumerate(header_row)]
                data: list[list[Any]] = []
                n = 0
                for row in reader:
                    if max_data_rows is not None and n >= max_data_rows:
                        break
                    padded = list(row) + [""] * (len(headers) - len(row))
                    data.append(padded[: len(headers)])
                    n += 1
                return headers, data
        except UnicodeDecodeError as e:
            last_err = e
            continue
    raise ValueError(f"CSV 인코딩을 알 수 없습니다: {path.name}") from last_err


def _load_xlsx(
    path: Path,
    *,
    sheet_index: int,
    max_data_rows: int | None,
) -> tuple[list[str], list[list[Any]]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[sheet_index]
        it = ws.iter_rows(values_only=True)
        try:
            first = next(it)
        except StopIteration:
            return [], []
        headers = [_cell_str(c, i) for i, c in enumerate(first)]
        data: list[list[Any]] = []
        n = 0
        for row in it:
            if max_data_rows is not None and n >= max_data_rows:
                break
            row_list = list(row) if row is not None else []
            padded = list(row_list) + [None] * (len(headers) - len(row_list))
            data.append(padded[: len(headers)])
            n += 1
        return headers, data
    finally:
        wb.close()


def _load_xlsx_by_sheet_name(
    path: Path,
    *,
    sheet_name: str,
    max_data_rows: int | None,
) -> tuple[list[str], list[list[Any]]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"시트를 찾을 수 없습니다: {sheet_name!r} (통합문서: {path.name})")
        ws = wb[sheet_name]
        it = ws.iter_rows(values_only=True)
        try:
            first = next(it)
        except StopIteration:
            return [], []
        headers = [_cell_str(c, i) for i, c in enumerate(first)]
        data: list[list[Any]] = []
        n = 0
        for row in it:
            if max_data_rows is not None and n >= max_data_rows:
                break
            row_list = list(row) if row is not None else []
            padded = list(row_list) + [None] * (len(headers) - len(row_list))
            data.append(padded[: len(headers)])
            n += 1
        return headers, data
    finally:
        wb.close()


def resolve_column_index(headers: list[str], column_name: str) -> int:
    """
    콤보박스에서 고른 열 이름 → 실제 테이블 헤더 인덱스.
    정확 일치 후, 앞뒤 공백 일치, 마지막으로 대소문자 무시 일치(영문 헤더용).
    """
    if not column_name or not str(column_name).strip():
        raise KeyError("열 이름이 비어 있습니다")
    if column_name in headers:
        return headers.index(column_name)
    want = column_name.strip()
    matches = [i for i, h in enumerate(headers) if h.strip() == want]
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        return matches[0]
    want_lower = want.lower()
    ci_matches = [i for i, h in enumerate(headers) if h.strip().lower() == want_lower]
    if len(ci_matches) >= 1:
        return ci_matches[0]
    raise KeyError(
        f"열을 찾을 수 없습니다: {column_name!r} — 테이블 헤더: {list(headers)}"
    )


def values_for_column(headers: list[str], rows: list[list[Any]], column_name: str) -> list[Any]:
    """헤더 이름으로 열 값 전체."""
    idx = resolve_column_index(headers, column_name)
    out: list[Any] = []
    for row in rows:
        out.append(row[idx] if idx < len(row) else None)
    return out


def list_files_in_folder(folder: Path) -> list[Path]:
    folder = folder.resolve()
    if not folder.is_dir():
        return []
    exts = supported_extensions()
    files: list[Path] = []
    for p in sorted(folder.iterdir()):
        if p.is_file() and p.suffix.lower() in exts:
            files.append(p)
    return files

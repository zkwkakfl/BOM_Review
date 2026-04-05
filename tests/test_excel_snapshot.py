"""검토용 통합문서 스냅샷 기록."""

from pathlib import Path

from bom_review.excel_com import SelectionSourceMeta
from bom_review.excel_snapshot import (
    BOM_REVIEW_COPY_SHEET,
    RANGE_SET_SHEET,
    apply_review_selection_to_snapshot,
    destination_sheet_name_for_role,
    finalize_snapshot_openpyxl,
    new_snapshot_workbook_path,
    write_role_range_to_snapshot,
)
from bom_review.table_io import load_header_and_rows_by_sheet_name


def test_apply_review_keeps_source_file_from_phase1_meta(tmp_path: Path) -> None:
    """2단계 apply_review는 원본파일·원본시트를 1단계 메타에서 유지한다."""
    from openpyxl import Workbook, load_workbook

    p = tmp_path / "twostep.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = BOM_REVIEW_COPY_SHEET
    ws.append(["Part", "Ref"])
    ws.append(["A", "R1 R2"])
    ws.append(["B", "R3"])
    wb.save(p)
    wb.close()

    source_meta = SelectionSourceMeta("원본.xlsx", "BOM시트", "$A$1:$B$3", "$A$1:$B$3")
    dummy = SelectionSourceMeta("?", "?", "$A$1:$B$2", "$A$1:$B$2")
    full_h = ["Part", "Ref"]
    full_d = [["A", "R1 R2"], ["B", "R3"]]
    rev_h = list(full_h)
    rev_d = [["A", "R1 R2"]]
    parsed = (full_h, full_d, rev_h, rev_d, dummy, 1, 1, 1, 1, 2, 2)

    merged = apply_review_selection_to_snapshot(
        p,
        role="BOM",
        dest_sheet_name=BOM_REVIEW_COPY_SHEET,
        source_meta=source_meta,
        parsed=parsed,
        bom_coord_excel_col_1based=2,
    )
    assert merged.source_file == "원본.xlsx"
    assert merged.source_sheet == "BOM시트"
    assert merged.review_range_address == "$A$1:$B$2"

    wb2 = load_workbook(p)
    try:
        w = wb2[BOM_REVIEW_COPY_SHEET]
        assert w.cell(row=2, column=2).value == "R1, R2"
    finally:
        wb2.close()


def test_destination_sheet_name_for_role() -> None:
    assert destination_sheet_name_for_role("BOM") == BOM_REVIEW_COPY_SHEET
    assert destination_sheet_name_for_role("원본") == "원본_검토복사"


def test_write_two_roles_updates_range_set(tmp_path: Path) -> None:
    p = new_snapshot_workbook_path(tmp_path)
    meta_b = SelectionSourceMeta("bom.xlsx", "Sheet1", "$A$1:$B$5", "$A$1:$B$4")
    s1 = write_role_range_to_snapshot(
        p,
        role="BOM",
        headers=["Ref", "Qty"],
        data_rows=[["R1", 1]],
        meta=meta_b,
        create_new_workbook=True,
    )
    assert s1 == "BOM_검토복사"
    meta_s = SelectionSourceMeta("src.xlsx", "좌표", "$C$1:$C$9", "$C$2:$C$8")
    s2 = write_role_range_to_snapshot(
        p,
        role="원본",
        headers=["Ref"],
        data_rows=[["R1"]],
        meta=meta_s,
        create_new_workbook=True,
    )
    assert s2 == "원본_검토복사"

    h1, r1 = load_header_and_rows_by_sheet_name(p, sheet_name=s1, max_data_rows=None)
    assert "Ref" in h1
    assert r1[0][h1.index("Ref")] == "R1"

    from openpyxl import load_workbook

    wb = load_workbook(p)
    try:
        rs = wb[RANGE_SET_SHEET]
        assert rs.max_row == 3
        roles = [rs.cell(row=i, column=1).value for i in range(2, rs.max_row + 1)]
        assert "BOM" in roles
        assert "원본" in roles
    finally:
        wb.close()


def test_bom_snapshot_normalizes_coord_column_to_comma_space(tmp_path: Path) -> None:
    """BOM 복사 시 UsedRange 기준 좌표명 열 인덱스만 ', ' 형태로 통일."""
    p = new_snapshot_workbook_path(tmp_path)
    meta = SelectionSourceMeta("b.xlsx", "S", "$A$1:$C$3", "$A$1:$C$3")
    write_role_range_to_snapshot(
        p,
        role="BOM",
        headers=["Part", "Ref"],
        data_rows=[
            ["X", "R1 R2\tR3"],
            ["Y", "R4,R5"],
            ["Z", "R6;R7"],
        ],
        meta=meta,
        create_new_workbook=True,
        bom_coord_matrix_col_index=1,
    )
    h, rows = load_header_and_rows_by_sheet_name(p, sheet_name="BOM_검토복사", max_data_rows=None)
    ri = h.index("Ref")
    assert rows[0][ri] == "R1, R2, R3"
    assert rows[1][ri] == "R4, R5"
    assert rows[2][ri] == "R6, R7"
    pi = h.index("Part")
    assert rows[0][pi] == "X"


def test_finalize_skips_header_row_in_coord_column(tmp_path: Path) -> None:
    """헤더 셀은 공백이 구분자로 오해되지 않게 좌표 정규화에서 제외한다."""
    from openpyxl import Workbook, load_workbook

    p = tmp_path / "hdr.xlsx"
    sheet = BOM_REVIEW_COPY_SHEET
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(["자품목 번호", "위치정보"])
    ws.append(["P1", "R1 R2"])
    wb.save(p)
    wb.close()

    meta = SelectionSourceMeta("b.xlsx", "S", "$A$1:$B$2", "$A$1:$B$2")
    finalize_snapshot_openpyxl(
        p,
        role="BOM",
        dest_sheet_name=sheet,
        meta=meta,
        bom_coord_excel_col_1based=2,
        bom_norm_row_start=1,
        bom_norm_row_end=2,
        bom_first_data_row_1based=2,
    )

    wb2 = load_workbook(p)
    try:
        w = wb2[sheet]
        assert w.cell(row=1, column=1).value == "자품목 번호"
        assert w.cell(row=1, column=2).value == "위치정보"
        assert w.cell(row=2, column=2).value == "R1, R2"
    finally:
        wb2.close()


def test_finalize_skips_rows_until_after_review_header_row(tmp_path: Path) -> None:
    """UsedRange가 검토 헤더보다 위 행(제목 등)을 포함할 때, 헤더 행 공백이 깨지지 않는다."""
    from openpyxl import Workbook, load_workbook

    p = tmp_path / "title_hdr.xlsx"
    sheet = BOM_REVIEW_COPY_SHEET
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(["BOM 목록", None])
    ws.append(["자품목 번호", "위치정보"])
    ws.append(["P1", "R1 R2"])
    wb.save(p)
    wb.close()

    meta = SelectionSourceMeta("b.xlsx", "S", "$A$1:$B$3", "$A$2:$B$3")
    finalize_snapshot_openpyxl(
        p,
        role="BOM",
        dest_sheet_name=sheet,
        meta=meta,
        bom_coord_excel_col_1based=2,
        bom_norm_row_start=1,
        bom_norm_row_end=3,
        bom_first_data_row_1based=3,
    )

    wb2 = load_workbook(p)
    try:
        w = wb2[sheet]
        assert w.cell(row=2, column=1).value == "자품목 번호"
        assert w.cell(row=2, column=2).value == "위치정보"
        assert w.cell(row=3, column=2).value == "R1, R2"
    finally:
        wb2.close()


def test_finalize_only_normalizes_bom_review_copy_sheet_name(tmp_path: Path) -> None:
    """역할이 BOM이어도 시트명이 BOM 복사본이 아니면 좌표 열을 바꾸지 않는다."""
    from openpyxl import Workbook, load_workbook

    p = tmp_path / "src_only.xlsx"
    wrong_sheet = "원본_검토복사"
    wb = Workbook()
    ws = wb.active
    ws.title = wrong_sheet
    ws.append(["Ref"])
    ws.append(["R1 R2"])
    wb.save(p)
    wb.close()

    meta = SelectionSourceMeta("s.xlsx", "S", "$A$1:$A$2", "$A$1:$A$2")
    finalize_snapshot_openpyxl(
        p,
        role="BOM",
        dest_sheet_name=wrong_sheet,
        meta=meta,
        bom_coord_excel_col_1based=1,
        bom_norm_row_start=1,
        bom_norm_row_end=2,
    )

    wb2 = load_workbook(p)
    try:
        assert wb2[wrong_sheet].cell(row=2, column=1).value == "R1 R2"
    finally:
        wb2.close()


def test_finalize_snapshot_skips_merged_cells_in_bom_coord_column(tmp_path: Path) -> None:
    """병합된 좌표 열에서 MergedCell에 value 대입 시 read-only 오류가 나지 않아야 한다."""
    from openpyxl import Workbook, load_workbook

    p = tmp_path / "merged.xlsx"
    sheet = BOM_REVIEW_COPY_SHEET
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(["Part", "Ref"])
    ws.append(["A", "R1 R2"])
    ws.append(["B", "R5 R6"])
    ws.append(["C", None])
    ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
    wb.save(p)
    wb.close()

    meta = SelectionSourceMeta("b.xlsx", "S", "$A$1:$B$4", "$A$1:$B$4")
    finalize_snapshot_openpyxl(
        p,
        role="BOM",
        dest_sheet_name=sheet,
        meta=meta,
        bom_coord_excel_col_1based=2,
        bom_norm_row_start=1,
        bom_norm_row_end=4,
        bom_first_data_row_1based=2,
    )

    wb2 = load_workbook(p)
    try:
        w = wb2[sheet]
        assert w.cell(row=2, column=2).value == "R1, R2"
        assert w.cell(row=3, column=2).value == "R5, R6"
    finally:
        wb2.close()


def test_reapply_same_role_replaces_range_set_row(tmp_path: Path) -> None:
    p = new_snapshot_workbook_path(tmp_path)
    meta1 = SelectionSourceMeta("a.xlsx", "S", "$A$1:$A$2", "$A$1:$A$2")
    write_role_range_to_snapshot(
        p,
        role="BOM",
        headers=["H"],
        data_rows=[["1"]],
        meta=meta1,
        create_new_workbook=True,
    )
    meta2 = SelectionSourceMeta("a.xlsx", "S", "$B$1:$B$3", "$B$2:$B$2")
    write_role_range_to_snapshot(
        p,
        role="BOM",
        headers=["H"],
        data_rows=[["2"]],
        meta=meta2,
        create_new_workbook=True,
    )
    from openpyxl import load_workbook

    wb = load_workbook(p)
    try:
        rs = wb[RANGE_SET_SHEET]
        assert rs.max_row == 2
        assert rs.cell(row=2, column=4).value == meta2.sheet_used_range_address
        assert rs.cell(row=2, column=5).value == meta2.review_range_address
    finally:
        wb.close()
